"""Multi-tab Excel workbook for staff review (pandas + openpyxl formatting).

Risk columns answer "how cleanup-worthy is this report?"; duplicate columns
answer "could another report replace it?". Flags are advisory only. Nothing is
ever auto-deleted — every column is a recommendation.
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .clean import normalize_report_name, text
from .security import (
    mask_df,
    sanitize_df_for_excel,
    secure_mkdir,
    sensitive_mode_enabled,
)

# Per-recommendation review tabs (the Overall-Score tiers below 100; the 100 /
# hard-rule cases get their own "Hard Rule Matches" tab).
BANDS = [
    "High-Priority Decommissioning Review",
    "Decommissioning Review",
    "Owner Review / Monitor",
    "Keep",
]
BAND_FILL = {
    "Keep": "DCFCE7",
    "Owner Review / Monitor": "DBEAFE",
    "Decommissioning Review": "FEF3C7",
    "High-Priority Decommissioning Review": "FEE2E2",
}


def _fmt_date(v):
    if v is None or pd.isna(v):
        return ""
    return pd.Timestamp(v).strftime("%Y-%m-%d")


def _bool(v):
    if v is None or (not isinstance(v, str) and pd.isna(v)):
        return ""
    return "Yes" if bool(v) else "No"


def _int(v):
    return "" if v is None or pd.isna(v) else int(v)


def _pct(v):
    return "" if v is None or (not isinstance(v, (int, float))) or pd.isna(v) else v


def _flags(v):
    return "; ".join(v) if isinstance(v, (list, set, tuple)) else text(v)


def _list_str(v):
    if isinstance(v, (list, set, tuple)):
        return "; ".join(str(x) for x in v)
    return text(v)


def _reason_str(reasons) -> str:
    return "; ".join(
        f"{'+' if (rs.points or 0) >= 0 else ''}{rs.points} {rs.label}" if rs.points is not None
        else rs.label
        for rs in (reasons or [])
    )


def report_row(r: dict) -> dict:
    """One row for the All Reports tab."""
    trail = _reason_str(r.get("all_reasons", []))
    hard = r.get("is_hard_rule")
    return {
        "Report Name": text(r.get("report_name")),
        "Report Type": text(r.get("report_type")),
        "Owner": text(r.get("owner")),
        "Created By": text(r.get("created_by")),
        "Category": text(r.get("category")),
        "Report Tag": text(r.get("report_tag")),
        "Data Source": text(r.get("data_source")),
        "Worklet": _bool(r.get("worklet")),
        "Landing Page": text(r.get("landing_page")),
        "Areas Where Used": text(r.get("areas_used")),
        "Shared": _bool(r.get("shared")),
        "Report Fields": text(r.get("report_fields")),
        "Field Extraction Status": text(r.get("field_extraction_status")),
        "Created Date": _fmt_date(r.get("created_date")),
        "Last Updated Date": _fmt_date(r.get("last_updated")),
        "Last Run Date": _fmt_date(r.get("last_run_date")),
        "Effective Last Run Date": _fmt_date(r.get("effective_last_run_date")),
        "Comprehensive Exec Count": _int(r.get("comprehensive_exec_count")),
        "Runs Exec Count (6mo window)": _int(r.get("runs_exec_count")),
        "Number of Times Run": _int(r.get("times_run")),
        # ---- Overall Decommissioning Score (headline) ----
        "Overall Score": _int(r.get("overall_score")),
        "Recommendation": text(r.get("recommendation")),
        "Hard Rule Triggered": _bool(r.get("hard_rule_triggered")),
        "Hard Rule Name": text(r.get("hard_rule_name")),
        "Cleanup Risk Points": _int(r.get("cleanup_risk_points")),
        "Cleanup %": _int(r.get("cleanup_percentage")),
        "Business Protection Credit": _int(r.get("business_protection_credit")),
        "Cleanup Reasons": _reason_str(r.get("cleanup_reasons")),
        "Protection Reasons": _reason_str(r.get("protection_reasons")),
        # ---- Recurrence ----
        "Recurrence": text(r.get("recurrence_classification")),
        "Recurrence Cadence": text(r.get("recurrence_cadence")),
        "Distinct Requesters": _int(r.get("distinct_requesters")),
        # ---- Weighted duplicate evidence ----
        "Potential Duplicate": _bool(r.get("potential_duplicate")),
        "Potential Duplicate Of": text(r.get("potential_duplicate_of")),
        "Duplicate Similarity": _pct(r.get("duplicate_similarity")),
        "Duplicate Relationship": text(r.get("duplicate_relationship")),
        "Field Jaccard %": _pct(r.get("field_jaccard_similarity")),
        "Smaller Report Containment %": _pct(r.get("smaller_report_containment")),
        # ---- Legacy risk columns (cleanup sub-scores; kept for continuity) ----
        "usage_risk": "" if hard else r.get("usage_risk", ""),
        "age_risk": "" if hard else r.get("age_risk", ""),
        "usage_context_risk": "" if hard else r.get("usage_context_risk", ""),
        "total_risk_score": "" if hard else r.get("total_risk_score", ""),
        "recommendation": text(r.get("recommendation")),
        # ---- Flags (no score weight) ----
        "ownership_flags": _flags(r.get("ownership_flags")),
        "data_quality_flags": _flags(r.get("data_quality_flags")),
        "all_flags": _flags(r.get("all_flags")),
        # ---- Duplicate / consolidation ----
        "duplicate_group_id": text(r.get("dup_group_id")),
        "is_suggested_keeper": _bool(r.get("is_suggested_keeper")) if r.get("dup_group_id") else "",
        "suggested_keeper_report_name": text(r.get("suggested_keeper_report_name")),
        "candidate_similarity_score": _pct(r.get("candidate_similarity_score")),
        "report_name_similarity_percent": _pct(r.get("report_name_similarity_percent")),
        "field_count": _int(r.get("field_count")),
        "shared_field_count": _int(r.get("shared_field_count")),
        "field_similarity_percent": _pct(r.get("field_similarity_percent")),
        "field_containment_percent": _pct(r.get("field_containment_percent")),
        "missing_fields_compared_to_keeper": _flags(r.get("missing_fields_compared_to_keeper")),
        "extra_fields_in_keeper": _flags(r.get("extra_fields_in_keeper")),
        "duplicate_classification": text(r.get("duplicate_classification")),
        "suggested_action": text(r.get("suggested_action")),
        "consolidation_reason": text(r.get("consolidation_reason")),
        "Reason Trail": trail,
    }


# Column order for the focused decommission summary export.
SUMMARY_COLUMNS = ["Report Name", "Overall Score", "Recommendation", "Reason Trail",
                   "Duplicate Status", "Duplicate Group ID"]
DUPLICATE_STATUS = "Duplicate"
NOT_DUPLICATE_STATUS = "Not a Duplicate"


def decommission_summary_row(r: dict, duplicate_status: str = NOT_DUPLICATE_STATUS,
                             duplicate_group_id: str = "") -> dict:
    """One row for the focused decommission summary export: the original report
    name, the headline Overall Decommissioning Score, the recommendation, the
    reason trail that explains the score, and the duplicate status + group id.

    duplicate_status / duplicate_group_id are supplied by export_decommission_summary
    (computed across all records); a non-member's group id is the empty string, never
    None/NaN, so it exports as a blank cell.
    """
    return {
        "Report Name": text(r.get("report_name")),
        "Overall Score": _int(r.get("overall_score")),
        "Recommendation": text(r.get("recommendation")),
        "Reason Trail": _reason_str(r.get("all_reasons", [])),
        "Duplicate Status": duplicate_status,
        "Duplicate Group ID": duplicate_group_id,
    }


def assign_duplicate_group_ids(records: list[dict]) -> dict[str, str]:
    """Map each detected duplicate group to a deterministic public id (DG-001, ...).

    A report "belongs to a duplicate group" when the existing duplicate-detection
    logic stamped it with a ``dup_group_id`` (recommend.apply_duplicate_analysis) —
    that id already encodes the transitive connected component, so A-B + B-C land in
    one group here automatically. We renumber those internal ids into DG-00N by
    sorting the groups on the normalized alphabetical name of their first
    (alphabetically-earliest) report, so the same input always yields the same ids.
    """
    members_by_group: dict[str, list[str]] = {}
    for r in records:
        gid = r.get("dup_group_id")
        if gid:
            members_by_group.setdefault(gid, []).append(
                normalize_report_name(r.get("report_name")))
    # Order groups by their alphabetically-first member name; tie-break on the
    # internal id (itself deterministic) so numbering is stable.
    ordered = sorted(members_by_group.items(), key=lambda kv: (min(kv[1]), kv[0]))
    return {gid: f"DG-{i:03d}" for i, (gid, _names) in enumerate(ordered, start=1)}


def duplicate_identification(records: list[dict]) -> dict[int, tuple[str, str, tuple]]:
    """Per-report duplicate identification, keyed by report_uid:
    ``(Duplicate Status, Duplicate Group ID, sort_key)``.

    Single source of truth shared by the Excel summary export AND the Streamlit app,
    so the on-screen table, its CSV download, and the summary workbook show identical
    values and ordering. The sort_key places duplicate groups first (by DG id, then
    normalized name) and non-duplicates last (by normalized name); it is for ordering
    only and is never displayed.
    """
    dg_by_group = assign_duplicate_group_ids(records)
    out: dict[int, tuple[str, str, tuple]] = {}
    for r in records:
        gid = r.get("dup_group_id")
        dg = dg_by_group.get(gid, "") if gid else ""
        status = DUPLICATE_STATUS if dg else NOT_DUPLICATE_STATUS
        norm = normalize_report_name(r.get("report_name"))
        out[r["report_uid"]] = (status, dg, (0 if dg else 1, dg, norm))
    return out


def _autoformat(ws):
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 50)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")


def export_workbook(
    records: list[dict],
    groups,
    run_meta: dict,
    out_dir: str | Path,
    field_rollup_result: dict | None = None,
) -> Path:
    # [SECURITY] Owner-only perms on the output dir (holds HR data workbooks).
    out_dir = secure_mkdir(out_dir)
    path = out_dir / f"report-cleanup-{datetime.now():%Y-%m-%d}.xlsx"

    # Build each report's export row ONCE and reuse it across every tab (All Reports,
    # per-recommendation tiers, Hard Rule, Meta-Similar) — report_row is non-trivial
    # and was previously recomputed per tab.
    row_by_uid = {r["report_uid"]: report_row(r) for r in records}
    all_rows = [row_by_uid[r["report_uid"]] for r in records]
    df_all = pd.DataFrame(all_rows)
    by_uid = {r["report_uid"]: r for r in records}

    def band_is(label):
        return [row_by_uid[r["report_uid"]] for r in records
                if not r.get("is_hard_rule") and r.get("recommendation") == label]

    hard = [row_by_uid[r["report_uid"]] for r in records if r.get("is_hard_rule")]

    # Duplicate Groups tab (keeper first within each group).
    dup_rows = []
    for g in sorted(groups, key=lambda x: x.group_id):
        members = sorted(g.members, key=lambda u: (by_uid[u]["report_uid"] != g.keeper_uid, u))
        for u in members:
            r = by_uid[u]
            dup_rows.append({
                "Duplicate Group ID": g.group_id,
                "Keeper?": "KEEPER" if u == g.keeper_uid else "",
                "Report Name": text(r.get("report_name")),
                "Owner": text(r.get("owner")),
                "Data Source": text(r.get("data_source")),
                "Field Count": _int(r.get("field_count")),
                "Field Extraction Status": text(r.get("field_extraction_status")),
                "Field Similarity %": _pct(r.get("field_similarity_percent")),
                "Field Containment %": _pct(r.get("field_containment_percent")),
                "Classification": text(r.get("duplicate_classification")),
                "Suggested Action": text(r.get("suggested_action")),
                "Consolidation Reason": text(r.get("consolidation_reason")),
            })

    # Reason Detail tab.
    reason_rows = []
    for r in records:
        for rs in r.get("all_reasons", []):
            reason_rows.append({
                "Report Name": text(r.get("report_name")),
                "Category": rs.category,
                "Reason": rs.label,
                "Points": "" if rs.points is None else rs.points,
            })

    # Metadata Similar – Fields Unavailable tab.
    meta_similar_rows = [
        row_by_uid[r["report_uid"]] for r in records
        if r.get("duplicate_classification") == "Metadata Similar - Fields Unavailable"
    ]

    # Field Rollup tab.
    field_rollup_rows = []
    if field_rollup_result:
        for rk, entry in sorted((field_rollup_result.get("rollup") or {}).items()):
            field_rollup_rows.append({
                "Report Key": rk,
                "Field Count": entry.get("field_count", 0),
                "Report Fields": entry.get("field_list_text", ""),
                "Business Objects Used": "; ".join(entry.get("business_objects_used") or []),
                "Domains Used": "; ".join(entry.get("domains_used") or []),
                "Categories Used": "; ".join(entry.get("categories_used") or []),
            })

    # Field Extraction Diagnostics tab.
    field_diag = run_meta.get("field_diag", {})
    field_diag_rows = [{"Metric": k, "Value": str(v)} for k, v in field_diag.items()]

    # Unmatched Field Rows tab.
    unmatched_rows = []
    if field_rollup_result:
        for row in (field_rollup_result.get("unmatched_rows") or []):
            unmatched_rows.append({
                k: str(v) for k, v in row.items() if not k.startswith("_")
            } | {"_reason": row.get("_reason", "")})

    # Data Quality tab.
    cov = run_meta.get("field_coverage", {})
    dq_rows = [{"Field": k, "Percent Populated": v} for k, v in sorted(cov.items())]
    diag = run_meta.get("diag", {})
    dq_rows += [
        {"Field": "-- join: composite matches", "Percent Populated": diag.get("composite", 0)},
        {"Field": "-- join: name-only fallback", "Percent Populated": diag.get("name_only", 0)},
        {"Field": "-- join: unmatched (no exec)", "Percent Populated": diag.get("unmatched", 0)},
        {"Field": "-- join: ambiguous name", "Percent Populated": diag.get("ambiguous", 0)},
    ]
    for w in run_meta.get("warnings", []):
        dq_rows.append({"Field": "WARNING", "Percent Populated": w})

    # Config Snapshot tab.
    config_snapshot = run_meta.get("config_snapshot", "{}")
    try:
        config_dict = json.loads(config_snapshot) if isinstance(config_snapshot, str) else config_snapshot
        config_rows = _flatten_config(config_dict)
    except Exception:
        config_rows = [{"Key": "config_snapshot", "Value": str(config_snapshot)}]

    # Summary.
    counts = _counts(records, groups, field_diag)
    summary = [{"Metric": k, "Value": v} for k, v in counts.items()]
    summary.append({"Metric": "Generated", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        _prep(pd.DataFrame(summary)).to_excel(xw, sheet_name="Summary Dashboard", index=False)
        _prep(df_all).to_excel(xw, sheet_name="All Reports", index=False)
        _sheet(xw, "Hard Rule Matches", hard)
        for band in BANDS:
            _sheet(xw, band, band_is(band))
        _sheet(xw, "Duplicate Groups", dup_rows)
        _sheet(xw, "Reason Detail", reason_rows)
        _sheet(xw, "Meta Similar No Fields", meta_similar_rows)
        _sheet(xw, "Field Rollup", field_rollup_rows)
        _sheet(xw, "Field Extraction Diag", field_diag_rows)
        _sheet(xw, "Unmatched Field Rows", unmatched_rows)
        _sheet(xw, "Data Quality", dq_rows)
        _sheet(xw, "Config Snapshot", config_rows)

        for ws in xw.book.worksheets:
            _autoformat(ws)
        _highlight_keepers(xw.book["Duplicate Groups"])

    return path


def export_decommission_summary(records: list[dict], out_dir: str | Path) -> Path:
    """Write a focused single-sheet workbook with exactly four columns: original
    report name, Overall Decommissioning Score, recommendation, and reason trail.

    A stakeholder-facing summary distinct from the full export_workbook diagnostics.
    Rows are grouped for review: every duplicate group appears as a contiguous block
    (DG-001, DG-002, ...), members sorted by normalized name within the block, then
    all non-duplicates by normalized name. Reuses the same presentation-layer security
    (sensitive masking + Excel-injection sanitizing) as every other sheet.
    """
    # [SECURITY] Owner-only perms on the output dir (holds HR data workbooks).
    out_dir = secure_mkdir(out_dir)
    path = out_dir / f"decommission-summary-{datetime.now():%Y-%m-%d}.xlsx"

    ident = duplicate_identification(records)
    # Sort key (local only — never written) groups duplicates into contiguous blocks.
    keyed = sorted(
        ((ident[r["report_uid"]][2],
          decommission_summary_row(r, ident[r["report_uid"]][0], ident[r["report_uid"]][1]))
         for r in records),
        key=lambda kr: kr[0],
    )
    # Force column order even when records is empty (header-only sheet).
    df = pd.DataFrame([row for _key, row in keyed], columns=SUMMARY_COLUMNS)

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        _prep(df).to_excel(xw, sheet_name="Decommission Summary", index=False)
        ws = xw.book["Decommission Summary"]
        _autoformat(ws)
        _style_summary_sheet(ws)

    return path


# Fixed column widths (chars) for the summary; everything else stays readable.
_SUMMARY_WIDTHS = {
    "Report Name": 40, "Overall Score": 12, "Recommendation": 34,
    "Reason Trail": 70, "Duplicate Status": 16, "Duplicate Group ID": 16,
}


def _style_summary_sheet(ws):
    """Presentation polish for the decommission summary so it reads cleanly in Excel:
    a colored header band, sensible fixed column widths, centered scores/ids, and a
    WRAPPED Reason Trail so long trails are fully visible instead of clipped."""
    header_fill = PatternFill("solid", fgColor="1F4E78")     # dark blue band
    header_font = Font(bold=True, color="FFFFFF")
    headers = [c.value for c in ws[1]]

    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="left")

    for idx, name in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        if name in _SUMMARY_WIDTHS:
            ws.column_dimensions[letter].width = _SUMMARY_WIDTHS[name]
        if name == "Reason Trail":
            for (cell,) in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        elif name in ("Overall Score", "Duplicate Group ID", "Duplicate Status"):
            for (cell,) in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                cell.alignment = Alignment(horizontal="center", vertical="top")


def _prep(df: pd.DataFrame) -> pd.DataFrame:
    """Apply presentation-layer security to a frame before it is written.

    [SECURITY] Order matters: mask sensitive columns first (if sensitive mode
    is on), then neutralize Excel formula injection on every text cell.
    """
    if sensitive_mode_enabled():
        df = mask_df(df)
    return sanitize_df_for_excel(df)


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _safe_sheet_name(name: str) -> str:
    # Excel forbids \ / * ? : [ ] in sheet titles and caps length at 31 chars.
    return _INVALID_SHEET_CHARS.sub("-", str(name))[:31].strip() or "Sheet"


def _sheet(xw, name, rows):
    name = _safe_sheet_name(name)
    df = pd.DataFrame(rows) if rows else pd.DataFrame([{"Info": "No rows in this category"}])
    _prep(df).to_excel(xw, sheet_name=name, index=False)


def _highlight_keepers(ws):
    if ws.max_row < 2:
        return
    headers = [c.value for c in ws[1]]
    if "Keeper?" not in headers:
        return
    kcol = headers.index("Keeper?") + 1
    fill = PatternFill("solid", fgColor="DCFCE7")
    for row in ws.iter_rows(min_row=2):
        if row[kcol - 1].value == "KEEPER":
            for c in row:
                c.font = Font(bold=True)
                c.fill = fill


def _flatten_config(d: dict, prefix: str = "") -> list[dict]:
    """Recursively flatten a nested config dict into key-value rows."""
    rows = []
    for k, v in d.items():
        full_key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            rows.extend(_flatten_config(v, full_key))
        elif isinstance(v, list):
            rows.append({"Key": full_key, "Value": "; ".join(str(i) for i in v)})
        else:
            rows.append({"Key": full_key, "Value": str(v) if v is not None else ""})
    return rows


def _counts(records, groups, field_diag: dict | None = None) -> dict:
    def band_count(label):
        return sum(1 for r in records if not r.get("is_hard_rule") and r.get("recommendation") == label)

    c = {
        "Total reports analyzed": len(records),
        "Hard rule matches": sum(1 for r in records if r.get("is_hard_rule")),
    }
    for band in BANDS:
        c[f"{band} (band)"] = band_count(band)
    c["Duplicate groups"] = len(groups)
    c["Reports in duplicate groups"] = sum(len(g.members) for g in groups)
    c["Metadata similar - fields unavailable"] = sum(
        1 for r in records
        if r.get("duplicate_classification") == "Metadata Similar - Fields Unavailable"
    )
    if field_diag:
        c["Field export mode"] = field_diag.get("field_export_mode", "")
        c["Reports with fields matched"] = field_diag.get("total_reports_with_fields", 0)
        c["Reports without fields"] = field_diag.get("total_reports_without_fields", 0)
        c["Field join match rate (%)"] = field_diag.get("field_join_match_rate", 0)
    return c
